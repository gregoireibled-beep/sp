import os
import glob
from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app) # Permet aux pages locales de communiquer avec Flask

@app.route('/recuperer-historique', methods=['POST'])
def recuperer_historique():
    try:
        data = request.json
        filiere = data.get('filiere')
        
        # Le dossier par défaut utilisé par votre fichier SPC.html
        dossier_cible = f"W:/Consignes/DFN/Extrusion/SPC/DEFAUT/"
        
        # Si vous utilisez un sous-dossier spécifique par filière, ajustez ici :
        # dossier_cible = f"W:/Consignes/DFN/Extrusion/SPC/{filiere}/"

        if not os.path.exists(dossier_cible):
            return jsonify({"status": "success", "data": []})

        # On cherche tous les fichiers générés par la filière sélectionnée
        fichiers = glob.glob(os.path.join(dossier_cible, f"SPC_{filiere}_*.xlsx"))
        
        toutes_les_mesures = []
        for fichier in fichiers:
            try:
                wb = load_workbook(fichier, data_only=True)
                ws = wb.active
                
                # Lecture de la première ligne de données (Ligne 2)
                # Structure dictée par votre fonction enregistrerFinal()
                mesure = {
                    "filiere": ws.cell(row=2, column=1).value,
                    "code_article": ws.cell(row=2, column=2).value,
                    "of": ws.cell(row=2, column=3).value,
                    "operateur": ws.cell(row=2, column=4).value,
                    "date": ws.cell(row=2, column=5).value
                }
                toutes_les_mesures.append(mesure)
            except Exception as e:
                continue # Ignore un fichier corrompu ou ouvert
                
        return jsonify({"status": "success", "data": toutes_les_mesures}), 200
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

if __name__ == '__main__':
    app.run(port=5000)
