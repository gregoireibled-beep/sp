import os
import sys
import glob
import webbrowser
from datetime import datetime
from threading import Timer
from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook

# =====================================================================
# 1. CONFIGURATION DES CHEMINS SÉCURISÉE POUR LE .EXE
# =====================================================================

def obtenir_dossier_application():
    """Détermine le dossier réel contenant les fichiers HTML et JS"""
    if getattr(sys, 'frozen', False):
        # Si c'est l'exécutable compilé
        return os.path.dirname(sys.executable)
    else:
        # Si c'est le script de développement .py
        return os.path.dirname(os.path.abspath(__file__))

DOSSIER_APP = obtenir_dossier_application()

app = Flask(__name__)
CORS(app)


# =====================================================================
# 2. ROUTES DE NAVIGATION (Sert les pages HTML directement)
# =====================================================================

@app.route('/')
def page_principale():
    """Affiche la page de saisie SPC.html principale"""
    chemin_page = os.path.join(DOSSIER_APP, "SPC.html")
    
    if not os.path.exists(chemin_page):
        return f"Erreur 404 : Le fichier de saisie est introuvable au chemin : {chemin_page}", 404
    
    with open(chemin_page, 'r', encoding='utf-8') as f:
        return f.read()

@app.route('/historique_SPC.html')
def page_historique():
    """Affiche la page d'analyse historique_SPC.html"""
    chemin_page = os.path.join(DOSSIER_APP, "historique_SPC.html")
    
    if not os.path.exists(chemin_page):
        return f"Erreur 404 : Le fichier d'historique est introuvable au chemin : {chemin_page}", 404
        
    with open(chemin_page, 'r', encoding='utf-8') as f:
        return f.read()

@app.route('/articles.js')
def servir_javascript():
    """Permet aux pages HTML de charger le dictionnaire d'articles"""
    chemin_js = os.path.join(DOSSIER_APP, "articles.js")
    
    if not os.path.exists(chemin_js):
        return "/* Erreur 404 : Fichier articles.js introuvable */", 404
        
    with open(chemin_js, 'r', encoding='utf-8') as f:
        return f.read(), 200, {'Content-Type': 'application/javascript'}


# =====================================================================
# 3. ROUTES REQUÊTES DONNÉES (Sauvegarde et Lecture Excel)
# =====================================================================

@app.route('/enregistrer-spc', methods=['POST'])
def save_data():
    try:
        data = request.json
        dossier = data.get('dossier')
        nom_fichier = data.get('nomFichier')
        mesures = data.get('mesures') 

        if not os.path.exists(dossier):
            os.makedirs(dossier)

        wb = Workbook()
        ws = wb.active

        # Entêtes fixes du fichier Excel
        colonnes_fixes = [
            "date", "machine", "designation", "code_article", "filiere", "of", 
            "operateur", "lot_matiere_vierge", "lot_matiere_broye", 
            "longueur_mm", "poids_kg", "colorimetrie_L", "colorimetrie_A", "colorimetrie_B", "observations"
        ]
        colonnes_dynamiques = [cle for cle in mesures.keys() if cle not in colonnes_fixes]
        toutes_les_colonnes = colonnes_fixes + colonnes_dynamiques
        
        ws.append(toutes_les_colonnes)
        valeurs = [mesures.get(colonne, "") for colonne in toutes_les_colonnes]
        ws.append(valeurs)

        chemin_complet = os.path.join(dossier, nom_fichier)
        wb.save(chemin_complet)

        print(f"✅ Fichier enregistré : {nom_fichier}")
        return jsonify({"status": "success"}), 200
    except Exception as e:
        print(f"❌ Erreur Enregistrement : {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/recuperer-historique', methods=['POST'])
def recuperer_historique():
    try:
        data = request.get_json()
        filiere_selectionnee = data.get('filiere')
        annee_selectionnee = data.get('annee')
        
        # Base du chemin réseau
        base_reseau = "W:/Consignes/DFN/Extrusion/SPC/Historique_SPC_Extrusion"

        # 1. CONSTRUCTION DYNAMIQUE DU CHEMIN (Gestion du mot-clé "TOUS")
        partie_annee = "*" if annee_selectionnee == "TOUS" else annee_selectionnee
        partie_filiere = "*" if filiere_selectionnee == "TOUS" else filiere_selectionnee

        # Construction du pattern de recherche global
        # Exemple si TOUS/TOUS : W:/Consignes/DFN/Extrusion/SPC/Historique_SPC_Extrusion/*/*/*.xls*
        pattern_recherche = os.path.join(base_reseau, partie_annee, partie_filiere, "*.xls*")
        
        # Normalisation des slashes pour Windows/Réseau
        pattern_recherche = pattern_recherche.replace('\\', '/')

        # Recherche de tous les fichiers Excel correspondants
        fichiers = glob.glob(pattern_recherche)
        
        toutes_les_mesures = []
        for fichier in fichiers:
            try:
                wb = load_workbook(fichier, data_only=True)
                ws = wb.active
                
                # Formatage sécurisé de la date
                val_date = ws.cell(row=2, column=1).value
                if isinstance(val_date, datetime):
                    str_date = val_date.strftime("%d/%m/%Y %H:%M")
                else:
                    str_date = str(val_date) if val_date is not None else ""

                def clean_val(val):
                    return str(val).strip() if val is not None else ""

                # 1. Base fixe commune à tous vos fichiers SPC
                mesure = {
                    "date": str_date,
                    "machine": clean_val(ws.cell(row=2, column=2).value),
                    "designation": clean_val(ws.cell(row=2, column=3).value),
                    "code_article": clean_val(ws.cell(row=2, column=4).value),
                    "filiere": clean_val(ws.cell(row=2, column=5).value),
                    "of": clean_val(ws.cell(row=2, column=6).value),
                    "operateur": clean_val(ws.cell(row=2, column=7).value),
                    "lot_matiere_vierge": clean_val(ws.cell(row=2, column=8).value),
                    "lot_matiere_broye": clean_val(ws.cell(row=2, column=9).value)
                }

                # 2. LECTURE DYNAMIQUE DES COTES (De la colonne 10 jusqu'à la fin du fichier)
                for col_idx in range(10, ws.max_column + 1):
                    nom_entete = ws.cell(row=1, column=col_idx).value
                    valeur_cote = ws.cell(row=2, column=col_idx).value
                    
                    if nom_entete:
                        clef_exacte = str(nom_entete).strip()
                        mesure[clef_exacte] = clean_val(valeur_cote)

                toutes_les_mesures.append(mesure)
            except Exception as e:
                print(f"⚠️ Impossible de lire le fichier {fichier} : {str(e)}")
                continue
                
        return jsonify({"status": "success", "data": toutes_les_mesures}), 200
    except Exception as e:
        print(f"❌ Erreur Lecture Historique : {str(e)}")
        return jsonify({"status": "error", "message": str(e)}), 500


# =====================================================================
# 4. ROUTE POUR LES IMAGES
# =====================================================================

@app.route('/images/<nom_image>')
def servir_images(nom_image):
    """Va chercher l'image de la filière directement sur le réseau d'usine W:"""
    dossier_images = "W:/Consignes/DFN/Extrusion/SPC/Image"
    nom_base = os.path.splitext(nom_image)[0]
    
    chemin_image = os.path.join(dossier_images, f"{nom_base}.jpg")
    
    if not os.path.exists(chemin_image):
        chemin_image = os.path.join(dossier_images, f"{nom_base}.jpeg")
        
    if os.path.exists(chemin_image):
        mimetype = "image/jpeg"
        with open(chemin_image, 'rb') as f:
            return f.read(), 200, {'Content-Type': mimetype}
            
    print(f"❌ Image de filière introuvable sur le réseau W: {nom_base}.jpg (ou .jpeg)")
    return f"Image introuvable dans le dossier {dossier_images}", 404


# =====================================================================
# 5. SCRIPT DE DÉMARRAGE
# =====================================================================

def ouvrir_navigateur():
    webbrowser.open("http://127.0.0.1:5000/")

if __name__ == "__main__":
    print("==================================================")
    print("        LANCEMENT DU SYSTÈME SPC EXTRUSION        ")
    print("==================================================")
    print(f"Dossier de l'application : {DOSSIER_APP}")
    
    Timer(1, ouvrir_navigateur).start()
    app.run(host='127.0.0.1', port=5000, threaded=True)
